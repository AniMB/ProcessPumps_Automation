def excel_file(table_data):
    import code12
    from openpyxl import Workbook
    from openpyxl.styles import  Border, Side, Alignment,  Font
    import gui
    import datetime
    wb= Workbook()
    ws=wb.active

    def wrap_text(cell_range):
        wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for row in cell_range:
            for cell in row:
                cell.alignment = wrap_alignment
                
                
    font = Font(name='Calibri',
                size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    #border style for all tabular data
    left=Side(border_style='thick',color='FF000000')
    right=Side(border_style='thick',color='FF000000')
    top=Side(border_style='thick',color='FF000000')
    bottom=Side(border_style='thick',color='FF000000')
    thick_border1=Border(top=top,right=right,left=left,bottom=bottom)



    #increasing width of cells so data can fit in in
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 15


    #gui={'val1': 40, 'val2': 'l', 'val3': 'jk' , 'val4': 'tytyyt'}
    value=gui.data_entry()
    ws.title=(str(datetime.date.today()))

    def print_between_keys(dictionary, start_key, end_key):
        keys = list(dictionary.keys())
        start_index = keys.index(start_key)
        end_index = keys.index(end_key)
        list1=[]
        paired=[]
        for key in keys [start_index:end_index+1]:
            value = dictionary[key]
            paired.extend([key,value])
            list1.append(tuple(paired))
            paired=[]
        return list1



    #MAIN HEADING FOR EXCEL FILE
    ws.merge_cells('A1:R1')
    cell = ws.cell(row=1,column=1)
    cell.value = 'PUMP PERFORMANCE TESTING REPORT'
    cell. alignment = Alignment(horizontal='center', vertical='center')
    range=ws['A1':'R1']
    for cell in range:
        for x in cell:
            x.border=thick_border1





    #TABLE 1 HEADER

    cell = ws.cell(row=3,column=1)
    ws.merge_cells('A3:B3')
    cell.value = 'PUMP DETAILS'
    cell. alignment = Alignment(horizontal='center', vertical='center')
    range=ws['A3':'B3']
    for cell in range:
        for x in cell:
            x.border=thick_border1
    wrap_text(ws['A4:A8'])

    #PRINTING DATA FOR TABLE 1
    row=4
    first=print_between_keys(value,'sl_no_','special_')
    for i in first:
        ws.cell(row=row, column=1, value=i[0])
        ws.cell(row=row, column=2, value=i[1])
        row += 1

    #giving borders to data in table 1
    range=ws['A4':'B8']
    for cell in range:
        for x in cell:
            x.border = thick_border1




    #TABLE 2 HEADER
    cell = ws.cell(row=3,column=5)
    ws.merge_cells('E3:F3')
    cell.value = 'PUMP OPERATION CONDITION'
    cell. alignment = Alignment(horizontal='center', vertical='center')
    range=ws['E3':'F3']
    for cell in range:
        for x in cell:
            x.border=thick_border1
    wrap_text(ws['E4:E11'])

    #PRINTING DATA FOR TABLE 2
    second=print_between_keys(value,'guageHeight_','head_')
    row=4
    for i in second:
        ws.cell(row=row, column=5, value=i[0])
        ws.cell(row=row, column=6, value=i[1])
        row+= 1

    #giving borders to data in table 2
    range=ws['E4':'G11']
    for cell in range:
        for x in cell:
            x.border=thick_border1




    #TABLE 3 HEADER
    cell = ws.cell(row=3,column=10)
    ws.merge_cells('J3:K3')
    cell.value = 'TESTING MOTOR DETAILS'
    cell. alignment = Alignment(horizontal='center', vertical='center')
    range=ws['J3':'K3']
    for cell in range:
        for x in cell:
            x.border=thick_border1
    wrap_text(ws['J4:J11'])
    #Abdhishay Murthy Nandula
        #16/05/2023
        
        
    #PRINTING DATA FOR TABLE 3
    third=print_between_keys(value,'make_','customer_WO_')
    row=4
    for i in third:
        ws.cell(row=row, column=10, value=i[0])
        ws.cell(row=row, column=11, value=i[1])
        row += 1

    #giving borders to data in table 3
    range=ws['J4':'K11']
    for cell in range:
        for x in cell:
            x.border=thick_border1


    

    # set the headers
    headers = [ "Sl.No.","Speed (RPM)", "Suction Pressure (bar)", "Delivery Gauge Reading (bar)", "Suction Gauge Reading (m)", "Delivery Gauge Reading (m)", "Flow Rate (m/hr)", "Velocity Head (m)", "Total Head (m)", "Voltage (Volts)", "Current (Amps)", "Power (kW)","Temperature (C)","Power Input (kW)",  "Flow Rate (m/hr)","Total head (m)", "Power (kW)", "Efficiency (%)"]
    column_index = 1
    for i, val in enumerate(headers):
        ws.cell(row=13, column=column_index + i, value=val)

    #Abdhishay Murthy Nandula
        #16/05/2023

    #ws.append(headers)

    # populate the data
    data=table_data
    # data = [
    #     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17],
    #     [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34],
    #     [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51]
    # ]



    for cell in ws[13]:
        for row in data:
            ws.append(row)
        break

    # set the border style
    thick_border = Border(left=Side(style='thick'), 
                            right=Side(style='thick'),
                            top=Side(style='thick'),
                            bottom=Side(style='thick')
                            )

    # set the alignment to wrap text
    
    wrap_text(ws['A13:R13'])
    #celval=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']
    # apply the border to each cell in the table
    range=ws['A13':'R18']
    for cell in range:
        for x in cell:
            x.border=thick_border
            #cell.alignment = wrap_alignment
            #for x in celval:
            # ws.column_dimensions["A"].width = 7
            # ws.column_dimensions["B"].width = 11
            # ws.column_dimensions["C"].width = 25
            # ws.column_dimensions["D"].width = 29
            # ws.column_dimensions["E"].width = 23
            # ws.column_dimensions["F"].width = 24
            # ws.column_dimensions["G"].width = 17
            # ws.column_dimensions["H"].width = 15
            # ws.column_dimensions["I"].width = 15
            # ws.column_dimensions["J"].width = 13
            # ws.column_dimensions["K"].width = 13
            # ws.column_dimensions["L"].width = 10
            # ws.column_dimensions["M"].width = 15
            # ws.column_dimensions["N"].width = 15
            # ws.column_dimensions["O"].width = 17
            # ws.column_dimensions["P"].width = 13
            # ws.column_dimensions["Q"].width = 11
            # ws.column_dimensions["R"].width = 12

    #code12.table_maker(table_data, value['Customer_Name_'],wb,ws)
   
   
    # SETTING THE DATE AS WELL AS A SPACE FOR START AND END TIME
    cell = ws.cell(row=4,column=14)
    cell.value = 'DATE'
    cell2 = ws.cell(row=4,column=15)
    cell2.value = str(datetime.date.today())
    ws.column_dimensions["N"].width = 14

    range=ws['N4':'O6']
    for cell in range:
        for x in cell:
            x.border=thick_border1

    cell3 = ws.cell(row=5,column=14)
    cell3.value = 'Started at : '
    cell4 = ws.cell(row=6,column=14)
    cell4.value = 'Stopped at : '


    # # VIBRATION VALUES
    # cell5 = ws.cell(row=5,column=18)
    # cell5.value = 'X'
    # cell6 = ws.cell(row=6,column=18)
    # cell6.value = 'Y'
    # cell44 = ws.cell(row=7,column=18)
    # cell44.value = 'Z'
    # cell7=ws.cell(row=4,column=18)
    # cell7.value = 'Vibration (rms-velocity)'

    print_between_keys(value,'vibx_','vibz_')
    row=4
    for i in second:
        ws.cell(row=row, column=18, value=i[0])
        ws.cell(row=row, column=18, value=i[1])
        row+= 1



    range=ws['P4':'Q7']
    for cell in range:
        for x in cell:
            x.border=thick_border1


    #PRINTING DATA FOR VIBRATION
    fifth=print_between_keys(value,'make_','customer_WO_')
    row=6
    for i in fifth:
        ws.cell(row=row, column=17, value=i[0])
        ws.cell(row=row, column=17, value=i[1])
        row += 1

    #APPROVAL
    cell8 = ws.cell(row=4,column=20)
    cell8.value = 'TESTED BY : '
    cell9 = ws.cell(row=5,column=20)
    cell9.value = 'WITNESSED BY : '


    range=ws['S4':'U5']
    for cell in range:
        for x in cell:
            x.border=thick_border1
            
    wb.save(f"{value['Customer_Name_']}.xlsx")
    wb.close()



    #  Varun Shankar Duddu
    #  18/05/2023
table_data = [
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17],
            [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34],
            [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51]
        ]
excel_file(table_data)


